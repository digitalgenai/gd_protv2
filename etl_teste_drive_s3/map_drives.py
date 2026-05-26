"""
Mapeamento de estrutura dos dois Drives
========================================
Gera um Excel com duas abas:
  - "Planilhas"  → estrutura completa da pasta de fornecedores
  - "Imagens"    → estrutura completa da pasta PRODUTOS

Uso:
    python map_drives.py
"""

import logging
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from google.oauth2 import service_account
from googleapiclient.discovery import build

from config import (
    CREDENTIALS_PATH,
    DRIVE_IMAGES_FOLDER_ID,
    DRIVE_SPREADSHEETS_FOLDER_ID,
)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]

OUTPUT_PATH = "mapa_drives.xlsx"


def authenticate():
    creds = service_account.Credentials.from_service_account_file(
        CREDENTIALS_PATH, scopes=SCOPES
    )
    return build("drive", "v3", credentials=creds)


def list_recursive(service, folder_id: str, path: str = "", depth: int = 0) -> list[dict]:
    items = []
    page_token = None
    while True:
        resp = service.files().list(
            q=f"'{folder_id}' in parents and trashed=false",
            fields="nextPageToken, files(id, name, mimeType, size)",
            pageToken=page_token,
            pageSize=1000,
            orderBy="name",
        ).execute()

        for f in resp.get("files", []):
            current_path = f"{path}/{f['name']}" if path else f["name"]
            is_folder = f["mimeType"] == "application/vnd.google-apps.folder"
            items.append({
                "id":       f["id"],
                "name":     f["name"],
                "path":     current_path,
                "depth":    depth,
                "type":     "Pasta" if is_folder else "Arquivo",
                "mimetype": f["mimeType"],
                "size_kb":  round(int(f.get("size", 0)) / 1024, 1) if f.get("size") else "",
            })
            if is_folder:
                items.extend(list_recursive(service, f["id"], current_path, depth + 1))

        page_token = resp.get("nextPageToken")
        if not page_token:
            break
    return items


HEADER_FILL = PatternFill("solid", fgColor="D3D1C7")
FOLDER_FILL = PatternFill("solid", fgColor="E6F1FB")
DEPTH_FILLS = [
    PatternFill("solid", fgColor="FFFFFF"),
    PatternFill("solid", fgColor="F5F5F0"),
    PatternFill("solid", fgColor="EEEEEA"),
    PatternFill("solid", fgColor="E8E8E3"),
]


def write_sheet(ws, items: list[dict], title: str):
    headers = ["caminho_completo", "nome", "tipo", "profundidade", "mime_type", "tamanho_kb", "id"]
    col_widths = [70, 45, 8, 12, 45, 14, 36]

    ws.append(headers)
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i)
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = col_widths[i - 1]
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    for item in items:
        ws.append([
            item["path"],
            item["name"],
            item["type"],
            item["depth"],
            item["mimetype"],
            item["size_kb"],
            item["id"],
        ])
        fill = FOLDER_FILL if item["type"] == "Pasta" else DEPTH_FILLS[min(item["depth"], 3)]
        font_bold = item["type"] == "Pasta"
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.fill = fill
            cell.font = Font(bold=font_bold, size=10)

    log.info(f"Aba '{title}': {len(items)} itens mapeados.")


def main():
    log.info("=== Mapeamento de estrutura dos Drives ===")
    service = authenticate()
    log.info("Autenticado.")

    wb = openpyxl.Workbook()

    log.info("Mapeando pasta de PLANILHAS...")
    ws1 = wb.active
    ws1.title = "Planilhas"
    planilhas = list_recursive(service, DRIVE_SPREADSHEETS_FOLDER_ID)
    write_sheet(ws1, planilhas, "Planilhas")

    log.info("Mapeando pasta de IMAGENS (PRODUTOS)...")
    ws2 = wb.create_sheet("Imagens")
    imagens = list_recursive(service, DRIVE_IMAGES_FOLDER_ID)
    write_sheet(ws2, imagens, "Imagens")

    wb.save(OUTPUT_PATH)
    log.info(f"Mapa salvo em: {OUTPUT_PATH}")
    log.info("Abra o arquivo e verifique os caminhos para ajustar os filtros do script principal.")


if __name__ == "__main__":
    main()
