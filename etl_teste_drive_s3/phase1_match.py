"""
FASE 1 — Extração, Matching e Relatório de Análise
===================================================
1. Autentica no Google Drive via service account
2. Varre a pasta de planilhas recursivamente, lendo apenas "Tabelas Atuais"
3. Extrai SKUs/nomes de produto de todas as abas de cada planilha
4. Varre a pasta PRODUTOS recursivamente e lista todas as imagens
5. Para cada imagem:
   - Extrai o código do fornecedor do nome do arquivo (ex: "ARBO1 - JAL.jpg" → JAL)
   - Se o código estiver mapeado em SUPPLIER_CODE_MAP com valor != None → mapeamento direto
   - Caso contrário → fuzzy matching contra todos os SKUs de todos os fornecedores
6. Gera relatorio_analise.xlsx com as associações para revisão humana

Uso:
    python phase1_match.py
"""

import io
import os
import re
import unicodedata
import logging
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import xlrd
from rapidfuzz import fuzz, process
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload

from config import (
    CREDENTIALS_PATH,
    DRIVE_IMAGES_FOLDER_ID,
    DRIVE_SPREADSHEETS_FOLDER_ID,
    SIMILARITY_THRESHOLD,
    HIGH_CONFIDENCE_THRESHOLD,
    MEDIUM_CONFIDENCE_THRESHOLD,
    AMBIGUITY_GAP,
    IMAGE_EXTENSIONS,
    SKU_COLUMN_NAMES,
    TABS_TO_SKIP_KEYWORDS,
    CURRENT_TABLE_KEYWORD,
    SUPPLIER_CODE_MAP,
    EXCLUIR,
    SUPPLIER_ALIASES,
    ANALYSIS_REPORT_PATH,
)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]


# ---------------------------------------------------------------------------
# Versionamento de relatórios
# ---------------------------------------------------------------------------

def get_versioned_path(base_path: str) -> tuple[str, str, int]:
    """
    A partir de 'relatorio_analise.xlsx', calcula:
      - versioned_path → 'relatorio_analise_v001.xlsx'  (próximo disponível)
      - latest_path    → 'relatorio_analise_LATEST.xlsx' (sempre sobrescrito)
      - version_num    → número inteiro da versão atual

    Procura arquivos existentes do padrão base_vNNN.xlsx para determinar o
    próximo número a usar.
    """
    p = Path(base_path)
    stem   = p.stem   # e.g. "relatorio_analise"
    suffix = p.suffix  # e.g. ".xlsx"
    parent = p.parent

    pattern = re.compile(rf"^{re.escape(stem)}_v(\d+){re.escape(suffix)}$", re.IGNORECASE)
    existing_versions = []
    for f in parent.glob(f"{stem}_v*{suffix}"):
        m = pattern.match(f.name)
        if m:
            existing_versions.append(int(m.group(1)))

    version_num   = (max(existing_versions) + 1) if existing_versions else 1
    versioned_path = str(parent / f"{stem}_v{version_num:03d}{suffix}")
    latest_path    = str(parent / f"{stem}_LATEST{suffix}")

    return versioned_path, latest_path, version_num


def load_previous_stats(latest_path: str) -> dict | None:
    """
    Lê o arquivo LATEST (se existir) e extrai os totais da aba 'Resumo da Versão'.
    Retorna um dict com as chaves: total, direto, alta, media, baixa, sem_match, version.
    Retorna None se o arquivo não existir ou não tiver a aba esperada.
    """
    p = Path(latest_path)
    if not p.exists():
        return None

    try:
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        if "Resumo da Versão" not in wb.sheetnames:
            wb.close()
            return None

        ws = wb["Resumo da Versão"]
        stats = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1] is not None:
                stats[str(row[0]).strip()] = row[1]
        wb.close()
        return stats if stats else None
    except Exception as e:
        log.warning(f"Não foi possível ler stats da versão anterior: {e}")
        return None

# MIME types aceitos para planilhas — todos em minúsculas (como o Drive retorna)
SPREADSHEET_MIMES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",  # .xlsx
    "application/vnd.ms-excel",                                            # .xls
    "application/vnd.ms-excel.sheet.macroenabled.12",                     # .xlsm (lowercase!)
    "application/vnd.google-apps.spreadsheet",                            # Google Sheets nativo
}


# ---------------------------------------------------------------------------
# Autenticação
# ---------------------------------------------------------------------------

def authenticate():
    creds = service_account.Credentials.from_service_account_file(
        CREDENTIALS_PATH, scopes=SCOPES
    )
    return build("drive", "v3", credentials=creds)


# ---------------------------------------------------------------------------
# Utilitários de texto
# ---------------------------------------------------------------------------

def normalize(text: str) -> str:
    """Remove acentos, converte para minúsculas e comprime espaços.

    Também remove conteúdo entre parênteses (descritores tipo "(ACESSÓRIO)"),
    quebra barras e colapsa quebras de linha — ruídos que prejudicam o fuzzy.
    """
    text = str(text).strip().lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")
    text = re.sub(r"\([^)]*\)", " ", text)          # remove "(...)"
    text = re.sub(r"[_\-/\\|]+", " ", text)          # separadores → espaço
    text = re.sub(r"\s+", " ", text)                  # colapsa espaços/quebras
    return text.strip()


# Valores que NÃO são SKU de produto (cabeçalhos/decoração comuns nas planilhas)
SKU_JUNK_VALUES = {
    "foto", "fotos", "picture", "pictures", "imagem", "imagens", "image",
    "copia", "copias", "modelo", "modelos", "produto", "produtos",
    "acessorio", "acessorios", "ref", "referencia", "descricao", "descrição",
    "description", "code", "coluna", "colecao", "dimensoes", "dimensions",
    "nome", "item", "itens", "codigo", "cod", "sku", "obs", "total",
}


def is_junk_sku(norm_value: str) -> bool:
    """True se o valor normalizado é claramente cabeçalho/decoração, não SKU."""
    n = norm_value.strip()
    if not n:
        return True
    if n in SKU_JUNK_VALUES:
        return True
    # valores tipo "foto / picture", "foto 01", "copia 2" etc.
    tokens = n.split()
    if tokens and all(t in SKU_JUNK_VALUES or t.isdigit() for t in tokens):
        return True
    return False


def is_skip_tab(tab_name: str) -> bool:
    norm = normalize(tab_name)
    return any(kw in norm for kw in TABS_TO_SKIP_KEYWORDS)


def is_valid_sku(value: str) -> bool:
    v = str(value).strip()
    if len(v) < 3:
        return False
    if re.match(r"^[\d\s.,]+$", v):
        return False
    if v.startswith("*"):                     # nota/instrução, não SKU
        return False
    if re.match(r"^[\d.,]+\s*[xX×ØøⱭ]", v):    # dimensão tipo "0,50 x 0,50" / "1,10 Ø"
        return False
    if len(normalize(v).split()) > 8:          # frase/descrição longa, não SKU
        return False
    if is_junk_sku(normalize(v)):
        return False
    return True


# ---------------------------------------------------------------------------
# Parse do nome da imagem
# ---------------------------------------------------------------------------

def parse_image_name(filename: str) -> tuple[str, str | None]:
    """
    Extrai (nome_produto_limpo, codigo_fornecedor) do nome do arquivo.

    Padrão esperado: "PRODUTO[N] - CODIGO.ext"
    Exemplos:
        "ARBO1 - JAL.jpg"            → ("ARBO", "JAL")
        "AIR SUSPENSO1 - JAL.jpg"    → ("AIR SUSPENSO", "JAL")
        "CAJU 1 - FEE.jpg"           → ("CAJU", "FEE")
        "CLASS 25 1 - GRH.avif"      → ("CLASS 25", "GRH")
        "BOWL1 - LINA BO BARDI - ARPER.jpeg" → ("BOWL", "ARPER")
        "foto_produto.jpg"           → ("foto produto", None)
    """
    stem = Path(filename).stem
    # Remove extensão extra que às vezes sobra no nome ("DFL.jpg", "MOO.jpeg")
    stem = re.sub(r"\.(jpe?g|png|gif|webp|bmp|tiff?|svg|avif)$", "", stem, flags=re.IGNORECASE)

    # Divide no último " - ". A parte final é o código se, depois de limpar
    # pontuação, tiver 2–6 letras (tolera ponto final, minúsculas e nomes curtos
    # como "MOROSO"/"BTZk"/"ART.").
    parts = stem.rsplit(" - ", 1)
    code = None
    product_raw = stem

    if len(parts) == 2:
        candidate = re.sub(r"[^A-Za-z&]", "", parts[1].strip())
        if 2 <= len(candidate) <= 6:
            code = candidate.upper()
            product_raw = parts[0]

    # Remove sufixos de variante/sequência ao final:
    #   "Turtle T04" → "Turtle"  (variante com letra+número, exige espaço antes
    #    para não quebrar modelos com número colado como "UP50")
    product = re.sub(r"\s+[A-Za-z]\d+$", "", product_raw)
    #   "ARBO1" → "ARBO" | "CAJU 1" → "CAJU" | "CLASS 25" → "CLASS"
    product = re.sub(r"\s*\d+$", "", product).strip()
    if not product:                      # se sobrou vazio, mantém o original
        product = product_raw.strip()

    return product, code


# ---------------------------------------------------------------------------
# Google Drive — listagem recursiva com caminho completo
# ---------------------------------------------------------------------------

def list_folder_recursive(service, folder_id: str, path: str = "") -> list[dict]:
    """
    Retorna lista de dicts de arquivos (não pastas) com caminho completo em 'path'.
    """
    items = []
    page_token = None

    while True:
        resp = service.files().list(
            q=f"'{folder_id}' in parents and trashed=false",
            fields="nextPageToken, files(id, name, mimeType)",
            pageToken=page_token,
            pageSize=1000,
        ).execute()

        for f in resp.get("files", []):
            current_path = f"{path}/{f['name']}" if path else f["name"]

            if f["mimeType"] == "application/vnd.google-apps.folder":
                items.extend(list_folder_recursive(service, f["id"], current_path))
            else:
                items.append({
                    "id":       f["id"],
                    "name":     f["name"],
                    "mimeType": f["mimeType"],
                    "path":     current_path,
                })

        page_token = resp.get("nextPageToken")
        if not page_token:
            break

    return items


# ---------------------------------------------------------------------------
# Download de arquivo do Drive em memória
# ---------------------------------------------------------------------------

def download_file(service, file_id: str, mime_type: str) -> bytes:
    if mime_type == "application/vnd.google-apps.spreadsheet":
        export_mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        request = service.files().export_media(fileId=file_id, mimeType=export_mime)
    else:
        request = service.files().get_media(fileId=file_id)

    buf = io.BytesIO()
    downloader = MediaIoBaseDownload(buf, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Extração de SKUs das planilhas
# ---------------------------------------------------------------------------

# Cabeçalhos PRIMÁRIOS de nome de produto (vencem sempre que tiverem conteúdo).
# Só "modelo"/"nome" — NÃO inclui "produto", senão "Descrição do Produto" (uma
# coluna de descrição) venceria a coluna "Modelo" na Feeling.
NAME_STRONG_TOKENS = {"modelo", "model", "nome", "name"}
# Cabeçalhos secundários — podem ser nome OU código dependendo do fornecedor
# (ex.: "Referência" é o nome na Sollos, mas é código no Brazil Contemporaneo).
# Entre eles, a escolha é pelo CONTEÚDO (qual coluna tem nomes de verdade).
NAME_SECONDARY_TOKENS = {
    "descricao", "description", "produto", "product", "item",
    "referencia", "ref", "cod", "codigo", "code", "sku", "colecao",
}
HEADER_SCAN_ROWS = 40     # cabeçalho pode estar fundo (Dona Flor: linha 13)
CONTENT_SCAN_ROWS = 400   # quantas linhas amostrar para avaliar o conteúdo da coluna


def _header_tokens(cell) -> set:
    return set(normalize(str(cell)).split()) if cell else set()


def _is_namelike(v) -> bool:
    """True se o valor parece NOME de produto (tem palavra), não código/número/bullet."""
    if not v or not isinstance(v, str):
        return False
    s = v.strip()
    if len(s) < 3:
        return False
    if re.match(r"^[\d.,]+\s*[xX×ØøⱭ]", s):   # dimensão
        return False
    return sum(c.isalpha() for c in s) >= 3    # ao menos uma palavra de verdade


def _namelike_count(rows: list, idx: int) -> int:
    n = 0
    for r in rows:
        if idx < len(r) and r[idx] is not None and _is_namelike(str(r[idx])):
            n += 1
    return n


def find_sku_column(rows_sample: list, is_xlsx: bool = True) -> int | None:
    """
    Identifica a coluna do NOME do produto combinando cabeçalho + CONTEÚDO.

    1) Coluna com cabeçalho FORTE (modelo/produto/nome) que tenha conteúdo de nome.
       (ex.: Feeling usa 'Modelo'.)
    2) Senão, entre as colunas com cabeçalho conhecido (descrição/referência/cód/…),
       escolhe a que MAIS tem valores parecidos com nome de produto.
       Isso resolve o ambíguo 'Referência': na Sollos ela tem nomes (vence a
       'Descrição' cheia de '•'); no Brazil Contemporaneo ela tem códigos numéricos
       (perde para a 'Descrição').
    3) Fallback: primeira coluna com strings longas.
    """
    scan = rows_sample[:HEADER_SCAN_ROWS]
    content = rows_sample[:CONTENT_SCAN_ROWS]

    strong_cols, known_cols = set(), set()
    for row in scan:
        for idx, cell in enumerate(row):
            toks = _header_tokens(cell)
            if not toks:
                continue
            if toks & NAME_STRONG_TOKENS:
                strong_cols.add(idx)
            if toks & (NAME_STRONG_TOKENS | NAME_SECONDARY_TOKENS):
                known_cols.add(idx)

    # 1) cabeçalho forte com conteúdo de nome
    cands = [(idx, _namelike_count(content, idx)) for idx in strong_cols]
    cands = [c for c in cands if c[1] > 0]
    if cands:
        return max(cands, key=lambda c: c[1])[0]

    # 2) qualquer cabeçalho conhecido → decide pelo conteúdo
    cands = [(idx, _namelike_count(content, idx)) for idx in known_cols]
    cands = [c for c in cands if c[1] > 0]
    if cands:
        return max(cands, key=lambda c: c[1])[0]

    # 3) fallback: primeira coluna com strings longas
    for row in scan:
        for idx, cell in enumerate(row):
            if cell and isinstance(cell, str) and len(cell.strip()) > 4:
                return idx

    return None


def parse_collection_name(cell) -> str | None:
    """Extrai o nome da coleção de um título de seção 'COLEÇÃO <NOME> - <designer>'.

    Planilhas como a Dona Flor organizam os produtos em seções por coleção, e o
    nome que aparece nas imagens é o da COLEÇÃO (ex.: 'ANA', 'PRAIA BRAVA'), não o
    tipo na coluna 'Descrição' ('Poltrona', 'Sofá'). Retorna o nome ou None.
    """
    if not cell or not isinstance(cell, str):
        return None
    m = re.match(r"(?i)\s*cole[çc][aã]o\s+(.+)", cell.strip())
    if not m:
        return None
    # corta no primeiro separador de descritor/designer
    name = re.split(r"\s*[-–/|]\s*", m.group(1))[0].strip()
    return name or None


def extract_skus_from_xlsx(data: bytes, supplier: str) -> list[tuple]:
    skus = []
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as e:
        log.warning(f"  Não foi possível abrir xlsx: {e}")
        return skus

    for sheet_name in wb.sheetnames:
        if is_skip_tab(sheet_name):
            continue

        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(min_row=1, max_row=500, values_only=True))
        if not all_rows:
            continue

        sku_col = find_sku_column(all_rows)

        count = 0
        col_count = 0
        for row in all_rows:
            # 1) nome da coluna identificada
            if sku_col is not None and sku_col < len(row) and row[sku_col]:
                val = str(row[sku_col]).strip()
                if is_valid_sku(val):
                    skus.append((normalize(val), val, supplier))
                    col_count += 1
            # 2) nome de coleção em títulos de seção (qualquer coluna)
            for cell in row:
                name = parse_collection_name(cell)
                if name and is_valid_sku(name):
                    skus.append((normalize(name), name, supplier))
                    count += 1

        if col_count or count:
            log.info(f"  Aba '{sheet_name}': {col_count} SKUs (col {sku_col}) + {count} coleções")

    wb.close()
    return skus


def extract_skus_from_xls(data: bytes, supplier: str) -> list[tuple]:
    skus = []
    try:
        wb = xlrd.open_workbook(file_contents=data)
    except Exception as e:
        log.warning(f"  Não foi possível abrir xls (talvez criptografado): {e}")
        return skus

    for sheet_name in wb.sheet_names():
        if is_skip_tab(sheet_name):
            continue

        ws = wb.sheet_by_name(sheet_name)
        sample = [
            [ws.cell_value(r, c) for c in range(ws.ncols)]
            for r in range(min(20, ws.nrows))
        ]
        sku_col = find_sku_column(sample)
        if sku_col is None:
            continue

        count = 0
        for r in range(1, ws.nrows):
            val = ws.cell_value(r, sku_col)
            if val and is_valid_sku(str(val)):
                skus.append((normalize(str(val)), str(val).strip(), supplier))
                count += 1

        if count:
            log.info(f"  Aba '{sheet_name}': {count} SKUs (col {sku_col})")

    return skus


# ---------------------------------------------------------------------------
# Carrega todas as planilhas e constrói o dicionário de SKUs
# ---------------------------------------------------------------------------

def canonical_supplier(folder_name: str, filename: str = "", full_path: str = "") -> str:
    """Resolve o fornecedor a partir da pasta + NOME DO ARQUIVO, via SUPPLIER_ALIASES.

    Importante: às vezes o fornecedor real está no NOME DO ARQUIVO, não na pasta —
    ex.: a linha da Roberta Banqueri vem no arquivo 'PIU MOBILE ROBETA BANQUERI...'
    dentro da pasta da PiuMobile. Por isso procuramos as chaves de alias como
    substring no (caminho + arquivo), da mais específica (mais longa) para a menos.
    """
    hay = normalize(f"{full_path} {filename}")
    for key in sorted(SUPPLIER_ALIASES, key=len, reverse=True):
        if key in hay:
            return SUPPLIER_ALIASES[key]

    # Fallback: nome da pasta (igual ou começa com uma chave de alias)
    base = re.sub(r"\s*-\s*\d+\s*$", "", folder_name).strip()
    key = base.lower()
    for prefix, canon in SUPPLIER_ALIASES.items():
        if key == prefix or key.startswith(prefix):
            return canon
    return base


def load_all_skus(service) -> list[tuple]:
    """Retorna lista de (sku_normalizado, sku_original, fornecedor)."""
    if not DRIVE_SPREADSHEETS_FOLDER_ID:
        log.error("DRIVE_SPREADSHEETS_FOLDER_ID não configurado no .env")
        return []

    log.info("Listando arquivos de planilhas no Drive...")
    all_files = list_folder_recursive(service, DRIVE_SPREADSHEETS_FOLDER_ID)
    log.info(f"Total de arquivos encontrados: {len(all_files)}")

    all_skus = []
    processed = set()

    for f in all_files:
        mime = f["mimeType"].lower()
        if mime not in SPREADSHEET_MIMES:
            continue

        path = f["path"]
        if CURRENT_TABLE_KEYWORD not in path.lower():
            continue

        if f["id"] in processed:
            continue
        processed.add(f["id"])

        # Extrai nome do fornecedor da primeira pasta do caminho (com alias)
        supplier = canonical_supplier(path.split("/")[0], f["name"], path)

        log.info(f"Processando [{supplier}]: {f['name']}")
        try:
            data = download_file(service, f["id"], f["mimeType"])
        except Exception as e:
            log.error(f"  Erro ao baixar {f['name']}: {e}")
            continue

        name_lower = f["name"].lower()
        if name_lower.endswith(".xls") and not name_lower.endswith(".xlsx"):
            skus = extract_skus_from_xls(data, supplier)
        else:
            skus = extract_skus_from_xlsx(data, supplier)

        all_skus.extend(skus)

    # Deduplica SKUs idênticos do mesmo fornecedor (linhas repetidas na planilha),
    # principal fonte de "falso ambíguo" no matching.
    seen = set()
    deduped = []
    for norm, original, supplier in all_skus:
        key = (norm, supplier)
        if key in seen:
            continue
        seen.add(key)
        deduped.append((norm, original, supplier))

    removed = len(all_skus) - len(deduped)
    log.info(f"SKUs totais extraídos: {len(all_skus)} | após dedup: {len(deduped)} (-{removed})")
    return deduped


# ---------------------------------------------------------------------------
# Carrega todas as imagens do Drive
# ---------------------------------------------------------------------------

def load_all_images(service) -> list[dict]:
    log.info("Listando imagens na pasta PRODUTOS...")
    all_files = list_folder_recursive(service, DRIVE_IMAGES_FOLDER_ID)

    images = []
    for f in all_files:
        ext = Path(f["name"]).suffix.lower()
        if ext not in IMAGE_EXTENSIONS:
            continue

        product, code = parse_image_name(f["name"])
        images.append({
            "id":           f["id"],
            "name":         f["name"],
            "path":         f["path"],
            "product_name": product,
            "product_norm": normalize(product),
            "supplier_code": code,
        })

    log.info(f"Total de imagens encontradas: {len(images)}")
    return images


# ---------------------------------------------------------------------------
# Matching
# ---------------------------------------------------------------------------

def confidence_label(score: int) -> str:
    if score >= HIGH_CONFIDENCE_THRESHOLD:
        return "Alta"
    if score >= MEDIUM_CONFIDENCE_THRESHOLD:
        return "Média"
    return "Baixa"


def _best_contained(product_tokens: set, product_norm: str, pool: list) -> dict | None:
    """Dentre os SKUs de `pool` (cada um {tokens, norm, original, supplier}),
    retorna o que CONTÉM o nome do produto (tokens do produto ⊆ tokens do SKU).
    Preferência: match exato; senão o SKU mais específico (mais tokens)."""
    cands = [c for c in pool if product_tokens and product_tokens <= c["tokens"]]
    if not cands:
        return None
    exatos = [c for c in cands if c["norm"] == product_norm]
    if exatos:
        return min(exatos, key=lambda c: len(c["original"]))
    # mais específico = mais tokens (desempate: nome mais longo)
    return max(cands, key=lambda c: (len(c["tokens"]), len(c["original"])))


def match_images_to_skus(images: list[dict], skus: list[tuple]) -> list[dict]:
    """
    MATCH POR CONTENÇÃO + DUPLA CONFIRMAÇÃO (nome do produto + código do fornecedor).

    O nome da imagem é o NÚCLEO do produto; na planilha o SKU costuma vir embrulhado
    com categoria ("BUFFET TECA") e/ou descritor ("CANDLE HOLDER FOGO"). Por isso o
    match é por CONTENÇÃO: os tokens do nome da imagem ⊆ tokens do SKU. Havendo vários,
    pega-se o mais específico.

      - confirmado : tem código → busca SÓ na planilha do fornecedor do código e acha
                     um SKU que contém o nome (dupla confirmação, pronto p/ S3).
      - nome       : sem código → nome contido em SKUs de um único fornecedor.
      - ambiguo    : sem código → nome contido em SKUs de 2+ fornecedores → revisão.
      - sem_sku    : nenhum SKU contém o nome → fornecedor = código (se houver).
      - excluido   : código marcado para exclusão.

    Obs.: com código, a busca nunca sai do fornecedor do código → não há "conflito".
    """
    if not skus:
        log.error("Lista de SKUs vazia.")
        return []

    # Pré-tokeniza os SKUs; índice por fornecedor + lista global (para imagens sem código)
    all_sku = []
    by_supplier: dict[str, list] = {}
    for norm, original, supplier in skus:
        if not norm:
            continue
        entry = {"tokens": set(norm.split()), "norm": norm, "original": original, "supplier": supplier}
        all_sku.append(entry)
        by_supplier.setdefault(supplier, []).append(entry)

    results = []
    total = len(images)
    counts: dict[str, int] = {}

    for i, img in enumerate(images, 1):
        if i % 200 == 0:
            log.info(f"  Matching {i}/{total}...")

        code          = img.get("supplier_code")
        product_norm  = img["product_norm"]
        product_tokens = set(product_norm.split())
        code_supplier = SUPPLIER_CODE_MAP.get(code) if code else None

        forn = sku = conf = method = second = atype = ""
        score = 0
        ambiguous = above = False

        if code_supplier == EXCLUIR:
            method, conf = "excluido", "Excluído"

        elif code_supplier:
            # COM CÓDIGO → procura só na planilha do fornecedor do código
            best = _best_contained(product_tokens, product_norm, by_supplier.get(code_supplier, []))
            if best:
                method, conf = "confirmado", "Confirmado"
                forn, sku, score, above = code_supplier, best["original"], 100, True
            else:
                method, conf = "sem_sku", "Sem SKU"
                forn, above = code_supplier, True

        else:
            # SEM CÓDIGO → procura em todas as planilhas
            cands = [c for c in all_sku if product_tokens and product_tokens <= c["tokens"]]
            suppliers = sorted({c["supplier"] for c in cands})
            if not cands:
                method, conf = "sem_sku", "Sem SKU"
            elif len(suppliers) == 1:
                best = _best_contained(product_tokens, product_norm, by_supplier[suppliers[0]])
                method, conf = "nome", "Nome"
                forn, sku, score, above = suppliers[0], best["original"], 100, True
            else:
                method, conf = "ambiguo", "Ambíguo"
                score, ambiguous, atype = 100, True, "fornecedor"
                vistos = {}
                for c in cands:
                    vistos.setdefault(c["supplier"], c["original"])
                second = "; ".join(f"{o} — {s}" for s, o in sorted(vistos.items()))

        counts[method] = counts.get(method, 0) + 1
        results.append({
            **img,
            "fornecedor":       forn,
            "sku_match":        sku,
            "score":            score,
            "confidence":       conf,
            "second_candidate": second,
            "second_score":     0,
            "ambiguous":        ambiguous,
            "ambiguity_type":   atype,
            "method":           method,
            "above_threshold":  above,
        })

    log.info("Matching concluído: " + " | ".join(f"{k}={v}" for k, v in sorted(counts.items())))
    return results


# ---------------------------------------------------------------------------
# Geração do relatório Excel
# ---------------------------------------------------------------------------

GREEN_FILL  = PatternFill("solid", fgColor="EAF3DE")
YELLOW_FILL = PatternFill("solid", fgColor="FAEEDA")
RED_FILL    = PatternFill("solid", fgColor="FCEBEB")
BLUE_FILL   = PatternFill("solid", fgColor="E6F1FB")
PURPLE_FILL = PatternFill("solid", fgColor="EDE7F6")  # mapeado sem planilha
GREY_FILL   = PatternFill("solid", fgColor="ECECEC")  # excluído
HEADER_FILL = PatternFill("solid", fgColor="D3D1C7")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def style_header(ws, row_num: int, col_count: int):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[row_num].height = 28


def row_fill_for(confidence: str, ambiguous: bool, method: str) -> PatternFill:
    return {
        "confirmado": GREEN_FILL,   # dupla confirmação
        "nome":       BLUE_FILL,    # só por nome
        "sem_sku":    PURPLE_FILL,  # fornecedor certo, sem SKU
        "conflito":   RED_FILL,     # revisão
        "ambiguo":    YELLOW_FILL,  # revisão
        "excluido":   GREY_FILL,
    }.get(method, RED_FILL)


def write_data_rows(ws, rows_data, headers):
    for r in rows_data:
        conf     = r.get("confidence", "Sem match")
        method   = r.get("method", "fuzzy")
        ambig    = r.get("ambiguous", False)
        fill     = row_fill_for(conf, ambig, method)
        above    = r.get("above_threshold", False)

        default_action = "CONFIRMAR" if method == "confirmado" else ""

        row = [
            r.get("id", ""),
            r["name"],
            r["path"],
            r.get("supplier_code", ""),
            r.get("method", ""),
            r.get("fornecedor", ""),
            r.get("sku_match", ""),
            r.get("score", 0),
            conf,
            r.get("second_candidate", ""),
            r.get("second_score", 0),
            "Sim" if ambig else "Não",
            r.get("ambiguity_type", ""),
            default_action,
        ]
        ws.append(row)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.fill = fill
            cell.font = Font(size=10)
            cell.alignment = Alignment(vertical="center")
            cell.border = THIN_BORDER


HEADERS = [
    "drive_id", "imagem", "caminho_drive", "codigo_fornecedor", "metodo",
    "fornecedor", "sku_match", "score", "confiança",
    "2º candidato", "score_2º", "ambíguo", "tipo_ambig", "ação",
]
COL_WIDTHS = [22, 28, 45, 16, 10, 22, 30, 8, 10, 40, 9, 9, 12, 14]


def setup_sheet_columns(ws):
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def generate_report(results: list[dict], output_path: str,
                    version_num: int = 1, prev_stats: dict | None = None):
    wb = openpyxl.Workbook()

    all_valid  = [r for r in results if r.get("above_threshold")]
    # "Para Revisão": casos que exigem decisão humana (conflito de fornecedor / ambíguo)
    for_review = [r for r in results if r.get("method") in ("conflito", "ambiguo")]
    # "Sem Match": sem colocação possível (sem SKU e sem fornecedor pelo código)
    no_match   = [r for r in results if r.get("method") == "sem_sku"
                  and not r.get("fornecedor")]

    # --- Aba 1: Análise Completa ---
    ws1 = wb.active
    ws1.title = "Análise Completa"
    ws1.append(HEADERS)
    style_header(ws1, 1, len(HEADERS))
    setup_sheet_columns(ws1)
    write_data_rows(ws1, sorted(results, key=lambda x: x.get("score", 0), reverse=True), HEADERS)

    # --- Aba 2: Para Revisão ---
    ws2 = wb.create_sheet("Para Revisão")
    ws2.append(HEADERS)
    style_header(ws2, 1, len(HEADERS))
    setup_sheet_columns(ws2)
    write_data_rows(ws2, sorted(for_review, key=lambda x: x.get("score", 0), reverse=True), HEADERS)

    # --- Aba 3: Sem Match ---
    ws3 = wb.create_sheet("Sem Match")
    no_match_headers = ["imagem", "caminho_drive", "codigo_fornecedor", "melhor_candidato", "score", "ação"]
    ws3.append(no_match_headers)
    style_header(ws3, 1, len(no_match_headers))
    for w, col in zip([30, 50, 16, 35, 8, 14], range(1, 7)):
        ws3.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    ws3.freeze_panes = "A2"
    for r in no_match:
        ws3.append([
            r["name"], r["path"], r.get("supplier_code", ""),
            r.get("sku_match", ""), r.get("score", 0), "",
        ])
        for col in range(1, 7):
            cell = ws3.cell(row=ws3.max_row, column=col)
            cell.fill = RED_FILL
            cell.font = Font(size=10)
            cell.border = THIN_BORDER
    ws3.freeze_panes = "A2"

    # --- Aba 4: Resumo por Fornecedor ---
    cats = ["confirmado", "nome", "sem_sku", "conflito", "ambiguo"]
    ws4 = wb.create_sheet("Resumo por Fornecedor")
    ws4.append(["fornecedor", "total"] + cats)
    style_header(ws4, 1, 2 + len(cats))
    for col in range(1, 3 + len(cats)):
        ws4.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 16
    ws4.freeze_panes = "A2"

    from collections import defaultdict
    summary = defaultdict(lambda: {"total": 0, **{c: 0 for c in cats}})
    for r in results:
        m = r.get("method")
        if m == "excluido":
            s = "(excluído)"
        else:
            s = r.get("fornecedor") or "Sem fornecedor"
        summary[s]["total"] += 1
        if m in cats:
            summary[s][m] += 1

    for supplier, data in sorted(summary.items()):
        ws4.append([supplier, data["total"]] + [data[c] for c in cats])
        for col in range(1, 3 + len(cats)):
            cell = ws4.cell(row=ws4.max_row, column=col)
            cell.font = Font(size=10)
            cell.border = THIN_BORDER

    # --- Aba 5: Legenda de Códigos ---
    ws5 = wb.create_sheet("Legenda Códigos")
    ws5.append(["codigo", "fornecedor_mapeado", "qtd_imagens", "status"])
    style_header(ws5, 1, 4)
    for col, w in zip(range(1, 5), [12, 28, 14, 22]):
        ws5.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    ws5.freeze_panes = "A2"

    from collections import Counter
    code_counts = Counter(r.get("supplier_code") for r in results if r.get("supplier_code"))
    from config import SUPPLIER_CODE_MAP as CODE_MAP
    for code, count in sorted(code_counts.items()):
        mapped = CODE_MAP.get(code)
        if mapped == EXCLUIR:
            display, status, fill = "Excluir", "Excluído", GREY_FILL
        elif mapped:
            display, status, fill = mapped, "Mapeado", GREEN_FILL
        else:
            display, status, fill = "", "Pendente — preencha o fornecedor", YELLOW_FILL
        ws5.append([code, display, count, status])
        for col in range(1, 5):
            cell = ws5.cell(row=ws5.max_row, column=col)
            cell.fill = fill
            cell.font = Font(size=10)
            cell.border = THIN_BORDER

    # --- Aba 6: Resumo da Versão ---
    ws6 = wb.create_sheet("Resumo da Versão")
    ws6.column_dimensions["A"].width = 30
    ws6.column_dimensions["B"].width = 18
    ws6.column_dimensions["C"].width = 20
    ws6.column_dimensions["D"].width = 30

    # Calcula stats desta execução
    def cnt(m): return sum(1 for r in results if r.get("method") == m)
    total       = len(results)
    confirmado  = cnt("confirmado")
    nome        = cnt("nome")
    sem_sku     = cnt("sem_sku")
    conflito    = cnt("conflito")
    ambiguo     = cnt("ambiguo")
    excluido    = cnt("excluido")
    # "prontos para o S3" = colocação válida (confirmado + nome + sem_sku com fornecedor)
    prontos     = sum(1 for r in results if r.get("above_threshold"))
    run_ts      = datetime.now().strftime("%Y-%m-%d %H:%M")

    # Cabeçalho
    ws6.append(["métrica", "valor_atual", "valor_anterior", "delta"])
    style_header(ws6, 1, 4)
    ws6.freeze_panes = "A2"

    def delta_str(label, current_val):
        if prev_stats is None:
            return "—"
        prev_val = prev_stats.get(label)
        if prev_val is None:
            return "—"
        try:
            diff = int(current_val) - int(prev_val)
            if diff > 0:
                return f"+{diff}"
            return str(diff)
        except (ValueError, TypeError):
            return "—"

    def delta_fill(label, current_val, good_direction="up"):
        """Colorir delta: verde se melhorou, vermelho se piorou."""
        if prev_stats is None:
            return None
        prev_val = prev_stats.get(label)
        if prev_val is None:
            return None
        try:
            diff = int(current_val) - int(prev_val)
            if diff == 0:
                return None
            improved = diff > 0 if good_direction == "up" else diff < 0
            return GREEN_FILL if improved else RED_FILL
        except (ValueError, TypeError):
            return None

    metrics = [
        ("versão",        "versao",     version_num, None),
        ("data / hora",   "data_hora",  run_ts,      None),
        ("total imagens", "total",      total,       None),
        ("confirmado (nome+código)", "confirmado", confirmado, "up"),
        ("só por nome",   "nome",       nome,        "up"),
        ("sem SKU",       "sem_sku",    sem_sku,     None),
        ("conflito (revisão)",  "conflito", conflito, "down"),
        ("ambíguo (revisão)",   "ambiguo",  ambiguo,  "down"),
        ("excluídas",     "excluido",   excluido,    None),
        ("prontos p/ S3 (válidos)", "prontos", prontos, "up"),
    ]

    for label, key, val, direction in metrics:
        prev_val = (prev_stats or {}).get(label, "—")
        d_str  = delta_str(label, val) if direction else "—"
        ws6.append([label, val, prev_val, d_str])

        row_idx = ws6.max_row
        # Colorir delta
        if direction:
            fill = delta_fill(label, val, direction)
            if fill:
                ws6.cell(row=row_idx, column=4).fill = fill

        for col in range(1, 5):
            cell = ws6.cell(row=row_idx, column=col)
            cell.font = Font(size=10)
            cell.border = THIN_BORDER

    wb.save(output_path)
    log.info(f"Relatório salvo em: {output_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    log.info("=== FASE 1 — Extração, Matching e Relatório ===")

    service = authenticate()
    log.info("Autenticado no Google Drive.")

    skus   = load_all_skus(service)
    images = load_all_images(service)

    if not skus:
        log.error("Nenhum SKU extraído. Verifique DRIVE_SPREADSHEETS_FOLDER_ID e as planilhas.")
        return
    if not images:
        log.error("Nenhuma imagem encontrada. Verifique DRIVE_IMAGES_FOLDER_ID.")
        return

    log.info(f"Iniciando matching: {len(images)} imagens × {len(skus)} SKUs...")
    results = match_images_to_skus(images, skus)

    # --- Versionamento ---
    versioned_path, latest_path, version_num = get_versioned_path(ANALYSIS_REPORT_PATH)
    prev_stats = load_previous_stats(latest_path)

    if prev_stats:
        log.info(f"Versão anterior encontrada (v{int(prev_stats.get('versao', 0)):03d}). Calculando delta...")
    else:
        log.info("Nenhuma versão anterior encontrada. Esta é a v001.")

    log.info(f"Salvando relatório como versão v{version_num:03d}...")
    generate_report(results, versioned_path, version_num=version_num, prev_stats=prev_stats)

    # Sobrescreve o LATEST com uma cópia
    import shutil
    shutil.copy2(versioned_path, latest_path)
    log.info(f"LATEST atualizado: {latest_path}")

    direct = sum(1 for r in results if r.get("method") == "direto")
    high   = sum(1 for r in results if r.get("confidence") == "Alta")
    above  = sum(1 for r in results if r.get("above_threshold"))
    log.info(f"Resultado: {above}/{len(results)} com match | {direct} mapeamento direto | {high} alta confiança")
    log.info(f"Relatório versionado : {versioned_path}")
    log.info(f"Relatório LATEST     : {latest_path}")


if __name__ == "__main__":
    main()
