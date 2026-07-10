"""
Diagnóstico: colisões de SKU entre fornecedores.

Pré-requisito para a abordagem de "match exato pelo nome": garante que nenhum
nome de SKU (normalizado) aparece em DOIS fornecedores diferentes. Se aparecer,
o match exato ficaria ambíguo.

Lê todas as planilhas via Drive (reaproveita a extração do phase1_match) e gera:
  - resumo no console
  - colisoes_sku.xlsx com duas abas: "Colisões" e "Resumo por Fornecedor"

Uso:
    python check_sku_collisions.py
"""

import collections

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from phase1_match import authenticate, load_all_skus, normalize, log

OUT = "colisoes_sku.xlsx"


def main():
    log.info("=== Checagem de colisões de SKU entre fornecedores ===")
    service = authenticate()
    skus = load_all_skus(service)  # lista de (norm, original, fornecedor) — já deduplicada
    if not skus:
        log.error("Nenhum SKU extraído. Verifique o acesso às planilhas.")
        return

    # norm -> { fornecedor: set(grafias originais) }
    index = collections.defaultdict(lambda: collections.defaultdict(set))
    for norm, original, supplier in skus:
        if norm:
            index[norm][supplier].add(original.replace("\n", " ").strip())

    # Colisão = mesmo nome normalizado em 2+ fornecedores diferentes
    colisoes = {n: forns for n, forns in index.items() if len(forns) >= 2}

    total_nomes = len(index)
    total_skus = len(skus)
    log.info(f"SKUs (após dedup): {total_skus} | nomes únicos: {total_nomes}")
    log.info(f"Nomes em MAIS DE UM fornecedor (colisões): {len(colisoes)}")

    # ---- console: lista as colisões ----
    if colisoes:
        print("\n=== COLISÕES (nome igual, fornecedores diferentes) ===")
        for norm in sorted(colisoes):
            forns = colisoes[norm]
            partes = "  |  ".join(
                f"{f}: {', '.join(sorted(g))}" for f, g in sorted(forns.items())
            )
            print(f"  '{norm}'  ->  {partes}")
    else:
        print("\n✅ Nenhuma colisão: cada nome de SKU pertence a um único fornecedor.")
        print("   A abordagem de match exato pelo nome é segura.")

    # ---- xlsx ----
    wb = openpyxl.Workbook()
    H = Font(bold=True, color="FFFFFF")
    HF = PatternFill("solid", fgColor="305496")

    def header(ws):
        for c in ws[1]:
            c.font = H
            c.fill = HF
            c.alignment = Alignment(vertical="center")
        ws.freeze_panes = "A2"

    ws1 = wb.active
    ws1.title = "Colisões"
    ws1.append(["nome_normalizado", "qtd_fornecedores", "fornecedores", "grafias_originais"])
    for norm in sorted(colisoes):
        forns = colisoes[norm]
        ws1.append([
            norm,
            len(forns),
            " | ".join(sorted(forns.keys())),
            " || ".join(f"{f}: {', '.join(sorted(g))}" for f, g in sorted(forns.items())),
        ])
    header(ws1)

    ws2 = wb.create_sheet("Resumo por Fornecedor")
    ws2.append(["fornecedor", "qtd_skus"])
    porforn = collections.Counter(s[2] for s in skus)
    for f, q in sorted(porforn.items(), key=lambda x: -x[1]):
        ws2.append([f, q])
    ws2.append(["TOTAL", total_skus])
    header(ws2)

    for ws in wb.worksheets:
        for col in range(1, ws.max_column + 1):
            mx = max((len(str(c.value)) for c in ws[get_column_letter(col)] if c.value is not None), default=10)
            ws.column_dimensions[get_column_letter(col)].width = min(mx + 2, 80)

    wb.save(OUT)
    log.info(f"Relatório salvo em: {OUT}")


if __name__ == "__main__":
    main()
