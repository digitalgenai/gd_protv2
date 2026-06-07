"""
FASE 2 — Upload para o S3 (com dry-run e manifesto incremental)
================================================================
Lê o relatório aprovado, pega só os CONFIRMADOS (nome + código batem),
deduplica imagens, agrupa por produto (fornecedor + SKU), sobe pro S3 em
   s3://<bucket>/<fornecedor-slug>/<sku-slug>/<seq>.<ext>
e grava um manifesto (manifesto_s3.csv) que liga Drive → S3 (e depois → PSQL).

Uso:
    python phase2_upload.py --dry-run   # simula: mede tamanho, NÃO sobe nada
    python phase2_upload.py             # sobe de verdade e grava o manifesto

Pré-requisitos:
    - relatório com a coluna 'drive_id' (rode a Fase 1 atualizada) — senão usa fallback por nome
    - Credenciais AWS no .env: AWS_ACCESS_KEY_ID, AWS_SECRET_ACCESS_KEY, S3_BUCKET_NAME, AWS_REGION
"""

import argparse
import csv
import io
import logging
import os
import re
import unicodedata
from datetime import datetime

import boto3
import openpyxl
from botocore.exceptions import ClientError
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload

from config import (
    CREDENTIALS_PATH,
    AWS_ACCESS_KEY_ID,
    AWS_SECRET_ACCESS_KEY,
    AWS_REGION,
    S3_BUCKET_NAME,
    ANALYSIS_REPORT_PATH,
)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]
MANIFEST_PATH = "manifesto_s3.csv"
MANIFEST_PREVIEW_PATH = "manifesto_s3_preview.csv"
MANIFEST_FIELDS = [
    "image_id", "arquivo_original", "caminho_drive",
    "fornecedor", "fornecedor_slug", "sku", "sku_slug", "codigo",
    "seq", "md5", "s3_key", "s3_url", "etag", "status", "versao", "enviado_em",
]

CONTENT_TYPES = {
    "jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png",
    "gif": "image/gif", "webp": "image/webp", "bmp": "image/bmp",
    "tiff": "image/tiff", "tif": "image/tiff", "svg": "image/svg+xml",
    "avif": "image/avif",
}


# ---------------------------------------------------------------------------
# Autenticação
# ---------------------------------------------------------------------------

def authenticate_google():
    creds = service_account.Credentials.from_service_account_file(CREDENTIALS_PATH, scopes=SCOPES)
    return build("drive", "v3", credentials=creds)


def authenticate_s3():
    return boto3.client(
        "s3",
        aws_access_key_id=AWS_ACCESS_KEY_ID,
        aws_secret_access_key=AWS_SECRET_ACCESS_KEY,
        region_name=AWS_REGION,
    )


# ---------------------------------------------------------------------------
# Slug / chave S3
# ---------------------------------------------------------------------------

def slugify(text: str) -> str:
    """Slug seguro (hífen): minúsculo, sem acento, remove (parênteses) e '•'."""
    text = str(text).strip().lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")
    text = re.sub(r"\([^)]*\)", " ", text)        # remove conteúdo entre parênteses
    text = re.sub(r"[^\w\s-]", " ", text)          # tira •, /, pontuação
    text = re.sub(r"[\s_]+", "-", text)
    text = re.sub(r"-+", "-", text)
    return text.strip("-")


def ext_of(filename: str) -> str:
    e = filename.rsplit(".", 1)[-1].lower() if "." in filename else "jpg"
    return "jpg" if e == "jpeg" else e


def parse_seq(filename: str) -> int:
    """Número da foto no nome (p/ ordenar). Ex.: 'DOTY2 - JAL.jpg' -> 2."""
    stem = filename.rsplit(".", 1)[0]
    stem = re.sub(r"\s*-\s*[A-Za-z.&]{2,8}$", "", stem)   # tira ' - CODIGO'
    m = re.search(r"(\d+)\s*$", stem)
    return int(m.group(1)) if m else 1


# ---------------------------------------------------------------------------
# Leitura do relatório
# ---------------------------------------------------------------------------

def resolve_report_path() -> str:
    latest = "relatorio_analise_LATEST.xlsx"
    return latest if os.path.exists(latest) else ANALYSIS_REPORT_PATH


def read_confirmed_rows(report_path: str) -> list[dict]:
    """Lê 'Análise Completa' e devolve só os confirmados (nome + código batem)."""
    wb = openpyxl.load_workbook(report_path, read_only=True, data_only=True)
    if "Análise Completa" not in wb.sheetnames:
        log.error("Aba 'Análise Completa' não encontrada.")
        return []
    ws = wb["Análise Completa"]
    headers, rows = [], []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c).strip().lower() if c else "" for c in row]
            continue
        if not any(row):
            continue
        rec = dict(zip(headers, row))
        metodo = str(rec.get("metodo", "")).strip().lower()
        acao   = str(rec.get("ação", "")).strip().upper()
        if metodo == "confirmado" or acao == "CONFIRMAR":
            rows.append(rec)
    wb.close()
    log.info(f"{len(rows)} linhas confirmadas no relatório.")
    return rows


def clean_sku(sku: str) -> str:
    return re.sub(r"\s+", " ", re.sub(r"\([^)]*\)", " ", str(sku))).strip()


def build_products(rows: list[dict]) -> tuple[list[dict], dict]:
    """Deduplica imagens e monta as linhas do manifesto (re-sequenciadas por produto)."""
    has_drive_id = any(str(r.get("drive_id", "")).strip() for r in rows)

    # dedup por drive_id (ideal) ou por caminho+arquivo (fallback)
    seen, dedup = set(), []
    for r in rows:
        fname = str(r.get("imagem", "")).strip()
        path  = str(r.get("caminho_drive", "")).strip()
        did   = str(r.get("drive_id", "")).strip()
        key = did or f"{path.lower()}|{fname.lower()}"
        if not fname or key in seen:
            continue
        seen.add(key)
        dedup.append({
            "image_id": did,
            "arquivo_original": fname,
            "caminho_drive": path,
            "fornecedor": str(r.get("fornecedor", "")).strip(),
            "sku": clean_sku(r.get("sku_match", "")),
            "codigo": str(r.get("codigo_fornecedor", "")).strip(),
        })

    # adiciona slugs (a sequência e o s3_key são definidos depois da dedup por md5)
    for d in dedup:
        d["fornecedor_slug"] = slugify(d["fornecedor"])
        d["sku_slug"] = slugify(d["sku"])

    stats = {
        "linhas_confirmadas": len(rows),
        "imagens_distintas": len(dedup),
        "produtos": len({(d["fornecedor_slug"], d["sku_slug"]) for d in dedup}),
        "tem_drive_id": has_drive_id,
    }
    return dedup, stats


# ---------------------------------------------------------------------------
# Drive helpers
# ---------------------------------------------------------------------------

def find_file_id_by_name(drive_service, filename: str) -> str | None:
    safe = filename.replace("'", "\\'")
    resp = drive_service.files().list(
        q=f"name='{safe}' and trashed=false",
        fields="files(id, name)", pageSize=5,
    ).execute()
    files = resp.get("files", [])
    return files[0]["id"] if files else None


def get_drive_meta(drive_service, file_id: str) -> tuple[int, str]:
    """Uma chamada de metadata: tamanho (bytes) + md5Checksum (hash de conteúdo)."""
    meta = drive_service.files().get(fileId=file_id, fields="size,md5Checksum").execute()
    return int(meta.get("size", 0) or 0), (meta.get("md5Checksum") or "")


def resolve_and_dedup(drive_service, items: list[dict]) -> tuple[list[dict], int, int, int]:
    """Resolve o file_id, busca size+md5, DEDUPLICA por conteúdo (md5) dentro de
    cada produto, re-sequencia 1..N e monta o s3_key.

    Retorna (final, total_bytes, sem_metadata, removidas_por_conteudo).
    """
    missing = 0
    for it in items:
        fid = it["image_id"] or find_file_id_by_name(drive_service, it["arquivo_original"])
        it["_file_id"] = fid or ""
        if not fid:
            it["_size"], it["md5"] = 0, ""
            missing += 1
            continue
        try:
            it["_size"], it["md5"] = get_drive_meta(drive_service, fid)
        except Exception as e:
            log.debug("metadata falhou p/ %s: %s", it["arquivo_original"], e)
            it["_size"], it["md5"] = 0, ""
            missing += 1

    # agrupa por produto e deduplica por NOME do arquivo dentro do produto:
    # o mesmo nome (DOTY1) repetido em várias pastas = a mesma foto re-salva —
    # mantém UMA cópia, a de MAIOR tamanho (melhor resolução). Re-sequencia 1..N.
    groups: dict = {}
    for it in items:
        groups.setdefault((it["fornecedor_slug"], it["sku_slug"]), []).append(it)

    final, removed = [], 0
    for (fslug, sslug), imgs in groups.items():
        # escolhe o representante de cada nome = maior arquivo
        por_nome: dict = {}
        for it in imgs:
            nome = it["arquivo_original"].strip().lower()
            atual = por_nome.get(nome)
            if atual is None:
                por_nome[nome] = it
            else:
                removed += 1
                # fica com o de maior tamanho
                if it.get("_size", 0) > atual.get("_size", 0):
                    por_nome[nome] = it
        kept = sorted(por_nome.values(),
                      key=lambda x: (parse_seq(x["arquivo_original"]), x["arquivo_original"].lower()))
        for seq, it in enumerate(kept, 1):
            it["seq"] = seq
            it["s3_key"] = f"{fslug}/{sslug}/{seq}.{ext_of(it['arquivo_original'])}"
            final.append(it)

    total_bytes = sum(it.get("_size", 0) for it in final)
    return final, total_bytes, missing, removed


def download_image(drive_service, file_id: str) -> bytes:
    request = drive_service.files().get_media(fileId=file_id)
    buf = io.BytesIO()
    downloader = MediaIoBaseDownload(buf, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    return buf.getvalue()


def s3_url(bucket: str, key: str) -> str:
    return f"https://{bucket}.s3.{AWS_REGION}.amazonaws.com/{key}"


def human(nbytes: int) -> str:
    for unit in ("B", "KB", "MB", "GB"):
        if nbytes < 1024:
            return f"{nbytes:.1f} {unit}"
        nbytes /= 1024
    return f"{nbytes:.1f} TB"


# ---------------------------------------------------------------------------
# Manifesto
# ---------------------------------------------------------------------------

def load_manifest(path: str) -> dict:
    if not os.path.exists(path):
        return {}
    out = {}
    with open(path, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            out[row.get("image_id") or row.get("s3_key")] = row
    return out


def write_manifest(path: str, records: list[dict]):
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=MANIFEST_FIELDS)
        w.writeheader()
        for r in records:
            w.writerow({k: r.get(k, "") for k in MANIFEST_FIELDS})


# ---------------------------------------------------------------------------
# Execução
# ---------------------------------------------------------------------------

def run(dry_run: bool):
    log.info("=== FASE 2 — Upload para S3 %s ===", "(DRY-RUN)" if dry_run else "")
    report = resolve_report_path()
    log.info("Relatório: %s", report)
    versao = re.search(r"(v\d+)", report)
    versao = versao.group(1) if versao else "LATEST"

    rows = read_confirmed_rows(report)
    if not rows:
        log.warning("Nenhuma linha confirmada. Nada a fazer.")
        return

    items, stats = build_products(rows)
    log.info("Confirmadas: %d | imagens (dedup por id): %d | produtos: %d",
             stats["linhas_confirmadas"], stats["imagens_distintas"], stats["produtos"])
    if not stats["tem_drive_id"]:
        log.warning("Relatório SEM coluna 'drive_id' — usando fallback por nome. "
                    "Rode a Fase 1 atualizada para dedup/idempotência exatas.")

    drive_service = authenticate_google()

    # Busca metadata (size+md5) e deduplica por NOME dentro do produto (mantém a
    # maior resolução) antes de qualquer upload
    log.info("Buscando metadata e deduplicando por nome (mantendo maior resolução)...")
    final, total_bytes, missing, removed = resolve_and_dedup(drive_service, items)
    log.info("Após dedup: %d imagens a subir (%d cópias removidas, %d sem metadata)",
             len(final), removed, missing)

    # ---- DRY-RUN: não sobe nada ----
    if dry_run:
        for it in final:
            it["s3_url"] = s3_url(S3_BUCKET_NAME or "<bucket>", it["s3_key"])
            it["status"] = "previsto"
            it["versao"] = versao
        write_manifest(MANIFEST_PREVIEW_PATH, final)
        log.info("=" * 56)
        log.info("DRY-RUN — nada foi enviado.")
        log.info("Imagens a subir: %d | produtos: %d", len(final), stats["produtos"])
        log.info("Cópias removidas (mesmo nome no produto): %d | sem metadata: %d", removed, missing)
        log.info("Tamanho total estimado: %s", human(total_bytes))
        log.info("Prévia do manifesto: %s", MANIFEST_PREVIEW_PATH)
        log.info("Custo de armazenamento S3 ≈ US$ %.4f/mês", (total_bytes / 1e9) * 0.023)
        return

    # ---- UPLOAD REAL ----
    if not S3_BUCKET_NAME:
        log.error("S3_BUCKET_NAME não configurado no .env.")
        return
    s3_client = authenticate_s3()
    manifest = load_manifest(MANIFEST_PATH)
    out_records, uploaded, skipped, errors = [], 0, 0, 0

    for i, it in enumerate(final, 1):
        key = it["s3_key"]
        it["s3_url"] = s3_url(S3_BUCKET_NAME, key)
        it["versao"] = versao
        prev = manifest.get(it["image_id"]) or manifest.get(key)

        # já no manifesto com o mesmo destino → pula
        if prev and prev.get("s3_key") == key and prev.get("status") == "enviado":
            it["status"], it["etag"] = "enviado", prev.get("etag", "")
            it["enviado_em"] = prev.get("enviado_em", "")
            out_records.append(it); skipped += 1
            continue

        fid = it.get("_file_id") or it["image_id"]
        if not fid:
            log.error("[%d/%d] não encontrado no Drive: %s", i, len(final), it["arquivo_original"])
            it["status"] = "erro"; out_records.append(it); errors += 1
            continue
        try:
            data = download_image(drive_service, fid)
            ext = ext_of(it["arquivo_original"])
            resp = s3_client.put_object(
                Bucket=S3_BUCKET_NAME, Key=key, Body=data,
                ContentType=CONTENT_TYPES.get(ext, "application/octet-stream"),
            )
            it["etag"] = (resp.get("ETag") or "").strip('"')
            it["status"] = "enviado"
            it["enviado_em"] = datetime.now().isoformat(timespec="seconds")
            out_records.append(it); uploaded += 1
            log.info("[%d/%d] %s → s3://%s/%s", i, len(final), it["arquivo_original"], S3_BUCKET_NAME, key)
        except Exception as e:
            log.error("[%d/%d] erro em %s: %s", i, len(final), it["arquivo_original"], e)
            it["status"] = "erro"; out_records.append(it); errors += 1

    write_manifest(MANIFEST_PATH, out_records)
    log.info("=" * 56)
    log.info("Concluído: %d enviados | %d já existiam | %d erros", uploaded, skipped, errors)
    log.info("Manifesto: %s", MANIFEST_PATH)


def main():
    ap = argparse.ArgumentParser(description="Upload de imagens confirmadas para o S3.")
    ap.add_argument("--dry-run", action="store_true", help="Simula: mede tamanho, não sobe nada.")
    args = ap.parse_args()
    run(dry_run=args.dry_run)


if __name__ == "__main__":
    main()
